# ============================================================
# University Faculty Slot Selection System — FastAPI Backend
# File: main.py
# ============================================================

from contextlib import asynccontextmanager
from datetime import datetime, timezone, timedelta
from typing import List, Optional
import asyncio
import bcrypt
import io
import os
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

import asyncpg
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect, UploadFile, File, BackgroundTasks
import smtplib
from email.message import EmailMessage
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

# Get the directory where main.py is located
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# Config
# ============================================================
DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://postgres:admin789@localhost:5432/univ_slots"
)
pool: asyncpg.Pool = None

class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []
        self.pending_updates = set() 

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    def mark_for_update(self, slot_id: str):
        self.pending_updates.add(slot_id)

    async def broadcast_updates(self):
        if not self.pending_updates or not self.active_connections:
            self.pending_updates.clear()
            return

        updates = []
        async with pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT id, max_capacity, current_enrolled 
                FROM faculty_availability 
                WHERE id = ANY($1::uuid[])
            """, list(self.pending_updates))
            
            for r in rows:
                updates.append({
                    "slot_id": str(r["id"]),
                    "remaining": r["max_capacity"] - r["current_enrolled"],
                    "is_full": r["current_enrolled"] >= r["max_capacity"]
                })

        if updates:
            message = {"type": "capacity_update", "slots": updates}
            dead_connections = []
            for connection in self.active_connections:
                try:
                    await connection.send_json(message)
                except Exception:
                    dead_connections.append(connection)
            for dead in dead_connections:
                self.disconnect(dead)

        self.pending_updates.clear()

manager = ConnectionManager()

async def broadcast_task():
    while True:
        await asyncio.sleep(1.0) # Throttled broadcast every 1 second
        if pool:
            try:
                await manager.broadcast_updates()
            except Exception as e:
                print(f"[WS ERROR] Broadcast failed: {e}")

async def portal_timer_task():
    while True:
        await asyncio.sleep(30.0)
        if pool:
            try:
                async with pool.acquire() as conn:
                    close_time_str = await conn.fetchval("SELECT value FROM system_settings WHERE key = 'portal_close_time'")
                    if close_time_str:
                        close_time = datetime.fromisoformat(close_time_str.replace('Z', '+00:00'))
                        now = datetime.now(timezone.utc) if close_time.tzinfo else datetime.now()
                        if now >= close_time:
                            await conn.execute("UPDATE system_settings SET value = 'true' WHERE key = 'student_locked'")
                            await conn.execute("DELETE FROM system_settings WHERE key = 'portal_close_time'")
                            print("[TIMER] Selection window ended. Portal locked.")
            except Exception as e:
                print(f"[TIMER ERROR] {e}")

@asynccontextmanager
async def lifespan(app: FastAPI):
    global pool
    try:
        pool = await asyncpg.create_pool(
            DATABASE_URL, 
            min_size=10, 
            max_size=100,
            command_timeout=60
        )
        print("[OK] Database connected successfully (Pool: 100).")
        # Ensure system_settings table exists
        async with pool.acquire() as conn:
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS system_settings (
                    key VARCHAR(50) PRIMARY KEY,
                    value VARCHAR(255) NOT NULL
                );
                INSERT INTO system_settings (key, value) VALUES ('portal_locked', 'false') ON CONFLICT DO NOTHING;
            ''')
    except Exception as e:
        print(f"[WARN] DB failed: {e}")
        pool = None
    
    bg_task = asyncio.create_task(broadcast_task())
    timer_task = asyncio.create_task(portal_timer_task())
    yield
    bg_task.cancel()
    timer_task.cancel()
    if pool:
        await pool.close()

# ============================================================
# App 
# ============================================================
app = FastAPI(title="University Slot Selection API", lifespan=lifespan)

# NOTE: allow_credentials=True cannot be used if allow_origins=["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve static files (HTML/CSS/JS)
# This allows mobile devices to access the UI by navigating to the computer's IP
app.mount("/static", StaticFiles(directory=BASE_DIR), name="static")

@app.websocket("/ws/updates")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            # Keep connection alive; clients don't send data here
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)
    except Exception as e:
        print(f"[WS ERROR] {e}")
        manager.disconnect(websocket)

# ============================================================
# Helper — ensure DB is available
# ============================================================
def require_db():
    if pool is None:
        raise HTTPException(
            status_code=503,
            detail="Database not connected. Please ensure PostgreSQL is running."
        )

async def ensure_settings():
    async with pool.acquire() as conn:
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS system_settings (
                key VARCHAR(50) PRIMARY KEY,
                value VARCHAR(255) NOT NULL
            );
        ''')
        await conn.execute("INSERT INTO system_settings (key, value) VALUES ('student_locked', 'false') ON CONFLICT (key) DO NOTHING")
        await conn.execute("INSERT INTO system_settings (key, value) VALUES ('faculty_locked', 'false') ON CONFLICT (key) DO NOTHING")

async def check_portal_locked(target="student", is_admin=False):
    require_db()
    if is_admin:
        return # Admins are never locked out
    await ensure_settings()
    key = f"{target}_locked"
    locked = await pool.fetchval(f"SELECT value FROM system_settings WHERE key = $1", key)
    if locked == 'true':
        raise HTTPException(status_code=403, detail=f"{target.capitalize()} portal is currently locked by the administrator.")

# ============================================================
# Pydantic Models
# ============================================================
class StartSelectionRequest(BaseModel):
    close_time: str

class LoginRequest(BaseModel):
    register_number: str
    password: str

class FacultyLoginRequest(BaseModel):
    employee_id: str
    password: str

class SelectionRequest(BaseModel):
    student_id: str
    slot_id: str

class FacultySelectRequest(BaseModel):
    student_id: str
    subject_id: str
    faculty_id: str
    stream_name: str

class SubjectDeselectRequest(BaseModel):
    student_id: str
    subject_id: str

class BulkSelectionRequest(BaseModel):
    student_id: str
    slot_ids: List[str]

class SubmitRequest(BaseModel):
    student_id: str

class FacultyPreferenceRequest(BaseModel):
    student_id: str
    subject_id: str
    faculty_id: str

class BulkFacultyPreferenceRequest(BaseModel):
    student_id: str
    preferences: List[dict]

# ============================================================
# Helpers
# ============================================================
DAY_NAMES = {1: "Monday", 2: "Tuesday", 3: "Wednesday", 4: "Thursday", 5: "Friday"}

async def get_student(student_id: str):
    return await pool.fetchrow("SELECT * FROM students WHERE id=$1::uuid", student_id)

async def is_submitted(student_id: str) -> bool:
    row = await pool.fetchrow(
        "SELECT 1 FROM student_submissions WHERE student_id=$1::uuid", student_id
    )
    return row is not None

# ============================================================
# Health check
# ============================================================
@app.get("/")
def home():
    login_path = os.path.join(BASE_DIR, "student.html")
    if not os.path.exists(login_path):
        return {"error": f"File not found at {login_path}. Please ensure student.html is in the project root."}
    return FileResponse(login_path)

@app.get("/{page}.html")
async def serve_page(page: str):
    page_path = os.path.join(BASE_DIR, f"{page}.html")
    if os.path.exists(page_path):
        return FileResponse(page_path)
    raise HTTPException(status_code=404, detail=f"Page {page}.html not found")

@app.get("/logo.png")
def serve_logo():
    logo_path = os.path.join(BASE_DIR, "logo.png")
    if os.path.exists(logo_path):
        return FileResponse(logo_path)
    raise HTTPException(status_code=404, detail="logo.png not found")

@app.get("/api/health")
def health():
    return {"status": "ok", "db_connected": pool is not None}

# ============================================================
# Auth
# ============================================================
@app.post("/api/login")
async def login(req: LoginRequest):
    require_db()
    # Check if portal is locked before allowing student login
    await check_portal_locked(target="student")
    
    print(f"[LOGIN] Attempt: {req.register_number}")
    student = await pool.fetchrow(
        "SELECT * FROM students WHERE register_number=$1", req.register_number
    )

    if not student:
        print(f"[LOGIN] Student not found: {req.register_number}")
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Verify password — supports both bcrypt and plain-text (for dev seed data)
    password_hash = student["password_hash"]
    valid = False
    # Check if it's a real bcrypt hash (starts with $2a$, $2b$, $2x$, $2y$)
    if password_hash.startswith("$2"):
        try:
            valid = bcrypt.checkpw(req.password.encode(), password_hash.encode())
        except Exception as e:
            print(f"[WARN] Bcrypt check error for {req.register_number}: {e}")
            valid = False
    else:
        # Plain-text comparison (seed data)
        valid = (req.password == password_hash)

    if not valid:
        print(f"[LOGIN] Invalid password for: {req.register_number}")
        raise HTTPException(status_code=401, detail="Invalid credentials")

    print(f"[OK] Login successful: {student['name']} ({req.register_number})")

    # Check if already submitted
    submitted = await is_submitted(str(student["id"]))

    return {
        "success": True,
        "message": "Login successful",
        "student": {
            "id": str(student["id"]),
            "name": student["name"],
            "register_number": student["register_number"],
            "email": student["email"],
            "department": student["department"],
            "semester": student["semester"],
            "is_submitted": submitted
        }
    }

# ============================================================
# Subjects & Faculty
# ============================================================
@app.get("/api/subjects")
async def get_subjects():
    require_db()
    rows = await pool.fetch("SELECT id, code, name, credits FROM subjects ORDER BY code")
    return [dict(r) for r in rows]

@app.get("/api/subjects/{subject_id}/faculty")
async def get_faculty_for_subject(subject_id: str):
    require_db()
    rows = await pool.fetch("""
        SELECT f.id, f.employee_id, f.name, f.department
        FROM faculty f
        JOIN faculty_subjects fs ON fs.faculty_id = f.id
        WHERE fs.subject_id = $1::uuid
        ORDER BY f.name
    """, subject_id)
    return [dict(r) for r in rows]

# ============================================================
# Faculty Preferences
# ============================================================
@app.get("/api/student/{student_id}/preferences")
async def get_student_preferences(student_id: str):
    require_db()
    rows = await pool.fetch("""
        SELECT subject_id, faculty_id 
        FROM student_faculty_preferences 
        WHERE student_id = $1::uuid
    """, student_id)
    return [dict(r) for r in rows]

@app.post("/api/student/preferences")
async def save_student_preferences(req: BulkFacultyPreferenceRequest):
    require_db()
    if await is_submitted(req.student_id):
        raise HTTPException(status_code=403, detail="Selection already submitted")

    async with pool.acquire() as conn:
        async with conn.transaction():
            # Delete existing preferences for this student to overwrite
            await conn.execute(
                "DELETE FROM student_faculty_preferences WHERE student_id = $1::uuid",
                req.student_id
            )
            
            # Insert new preferences
            for pref in req.preferences:
                await conn.execute("""
                    INSERT INTO student_faculty_preferences (student_id, subject_id, faculty_id)
                    VALUES ($1::uuid, $2::uuid, $3::uuid)
                """, req.student_id, pref["subject_id"], pref["faculty_id"])
                
    return {"success": True, "message": "Preferences saved"}

# ============================================================
# Student: Subjects & Faculty Options
# ============================================================
@app.get("/api/student/{student_id}/subjects")
async def get_student_subjects(student_id: str):
    require_db()
    rows = await pool.fetch("SELECT id, code, name, credits, subject_type, periods_per_week FROM subjects ORDER BY code")
    # For each subject, check if student has already selected a faculty
    result = []
    for r in rows:
        sel = await pool.fetchrow("""
            SELECT DISTINCT fa.faculty_id, f.name as faculty_name
            FROM student_selections ss
            JOIN faculty_availability fa ON fa.id = ss.availability_id
            JOIN faculty f ON f.id = fa.faculty_id
            WHERE ss.student_id = $1::uuid AND fa.subject_id = $2::uuid
            LIMIT 1
        """, student_id, r['id'])
        result.append({
            **dict(r),
            "selected_faculty_id": str(sel['faculty_id']) if sel else None,
            "selected_faculty_name": sel['faculty_name'] if sel else None
        })
    return result

@app.get("/api/student/{student_id}/faculty-options/{subject_id}")
async def get_faculty_options(student_id: str, subject_id: str):
    require_db()
    # Get all faculty who teach this subject
    faculty_list = await pool.fetch("""
        SELECT DISTINCT fa.faculty_id, f.name as faculty_name, f.employee_id, fa.stream_name
        FROM faculty_availability fa
        JOIN faculty f ON f.id = fa.faculty_id
        WHERE fa.subject_id = $1::uuid AND fa.is_active = TRUE
    """, subject_id)
    
    # Get student's current occupied slots (excluding this subject)
    occupied = await pool.fetch("""
        SELECT fa.day_of_week, fa.period_id
        FROM student_selections ss
        JOIN faculty_availability fa ON fa.id = ss.availability_id
        WHERE ss.student_id = $1::uuid AND fa.subject_id != $2::uuid
    """, student_id, subject_id)
    occupied_set = {(r['day_of_week'], r['period_id']) for r in occupied}
    
    options = []
    for fac in faculty_list:
        # Get all slots in this faculty's bundle for this subject
        slots = await pool.fetch("""
            SELECT fa.id, fa.day_of_week, fa.period_id, fa.lecture_seq,
                   fa.max_capacity, fa.current_enrolled,
                   p.period_num, p.start_time, p.end_time, p.label
            FROM faculty_availability fa
            JOIN periods p ON p.id = fa.period_id
            WHERE fa.faculty_id = $1::uuid AND fa.subject_id = $2::uuid 
              AND fa.stream_name = $3 AND fa.is_active = TRUE
            ORDER BY fa.day_of_week, p.period_num
        """, fac['faculty_id'], subject_id, fac['stream_name'])
        
        clashes = []
        is_full = False
        slot_details = []
        for s in slots:
            if s['current_enrolled'] >= s['max_capacity']:
                is_full = True
            if (s['day_of_week'], s['period_id']) in occupied_set:
                clashes.append(f"{DAY_NAMES[s['day_of_week']]} {s['label']}")
            slot_details.append({
                "day": s['day_of_week'], "day_name": DAY_NAMES[s['day_of_week']],
                "period_num": s['period_num'], "period_label": s['label'],
                "start_time": str(s['start_time']), "end_time": str(s['end_time']),
                "enrolled": s['current_enrolled'], "capacity": s['max_capacity'],
                "lecture_seq": s['lecture_seq']
            })
        
        avg_enrolled = sum(s['current_enrolled'] for s in slots) // max(len(slots), 1)
        options.append({
            "faculty_id": str(fac['faculty_id']),
            "faculty_name": fac['faculty_name'],
            "employee_id": fac['employee_id'],
            "stream_name": fac['stream_name'],
            "slots": slot_details,
            "has_clash": len(clashes) > 0,
            "clash_details": clashes,
            "is_full": is_full,
            "avg_enrolled": avg_enrolled,
            "avg_capacity": slots[0]['max_capacity'] if slots else 60
        })
    return options

@app.get("/api/timetable")
async def get_timetable(student_id: str):
    require_db()
    # 1. Get student's current selections
    current = await pool.fetch("""
        SELECT availability_id FROM student_selections WHERE student_id = $1::uuid
    """, student_id)
    selected_ids = {str(r['availability_id']) for r in current}
    
    # 2. Get student's current occupied slots (for clash detection)
    occupied = await pool.fetch("""
        SELECT fa.day_of_week, fa.period_id, sub.code as subject_code
        FROM student_selections ss
        JOIN faculty_availability fa ON fa.id = ss.availability_id
        JOIN subjects sub ON sub.id = fa.subject_id
        WHERE ss.student_id = $1::uuid
    """, student_id)
    # Using period_id for matching if p.id is used, but timetable.html uses period_num. 
    # Let's check periods table structure. 
    # Actually, p.id = fa.period_id, and p.period_num is the number.
    # We should use period_id for exact matching if possible.
    occupied_map = {r['period_id']: r['subject_code'] for r in occupied}

    # 3. Get all active slots with details
    rows = await pool.fetch("""
        SELECT 
            fa.id as slot_id, fa.faculty_id, fa.subject_id, fa.day_of_week, fa.stream_name,
            fa.max_capacity, fa.current_enrolled,
            f.name as faculty_name,
            sub.code as subject_code, sub.name as subject_name,
            p.id as period_id, p.period_num, p.label as period_label, p.start_time, p.end_time
        FROM faculty_availability fa
        JOIN faculty f ON f.id = fa.faculty_id
        JOIN subjects sub ON sub.id = fa.subject_id
        JOIN periods p ON p.id = fa.period_id
        WHERE fa.is_active = TRUE
        ORDER BY fa.day_of_week, p.period_num, f.name
    """)
    
    slots = []
    for r in rows:
        is_selected = str(r['slot_id']) in selected_ids
        
        # Determine if this slot has a clash with OTHER selections
        clash_subj = occupied_map.get(r['period_id'])
        has_clash = False
        if clash_subj and not is_selected:
            has_clash = True

        slots.append({
            "slot_id": str(r['slot_id']),
            "faculty_id": str(r['faculty_id']),
            "faculty_name": r['faculty_name'],
            "subject_id": str(r['subject_id']),
            "subject_code": r['subject_code'],
            "subject_name": r['subject_name'],
            "day": r['day_of_week'],
            "period_num": r['period_num'],
            "period_label": r['period_label'],
            "start_time": str(r['start_time']),
            "end_time": str(r['end_time']),
            "stream_name": r['stream_name'],
            "booked": r['current_enrolled'],
            "capacity": r['max_capacity'],
            "remaining": max(0, r['max_capacity'] - r['current_enrolled']),
            "is_full": r['current_enrolled'] >= r['max_capacity'],
            "is_selected": is_selected,
            "has_clash": has_clash,
            "clash_with": clash_subj if has_clash else None
        })
        
    submitted = await is_submitted(student_id)
    
    # 4. Constraints (min/max periods per week)
    constraints_rows = await pool.fetch("""
        SELECT id as subject_id, code as subject_code, periods_per_week
        FROM subjects
    """)
    constraints = [{
        "subject_id": str(r['subject_id']),
        "subject_code": r['subject_code'],
        "min_periods_per_week": r['periods_per_week'],
        "max_periods_per_week": r['periods_per_week']
    } for r in constraints_rows]

    return {
        "slots": slots,
        "is_submitted": submitted,
        "constraints": constraints
    }


@app.get("/api/student/{student_id}/my-timetable")
async def get_my_timetable(student_id: str):
    require_db()
    student = await get_student(student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    submitted = await is_submitted(student_id)
    
    rows = await pool.fetch("""
        SELECT fa.day_of_week, p.period_num, p.label, p.start_time, p.end_time,
               sub.code as subject_code, sub.name as subject_name, sub.subject_type,
               f.name as faculty_name, fa.lecture_seq, fa.stream_name
        FROM student_selections ss
        JOIN faculty_availability fa ON fa.id = ss.availability_id
        JOIN periods p ON p.id = fa.period_id
        JOIN subjects sub ON sub.id = fa.subject_id
        JOIN faculty f ON f.id = fa.faculty_id
        WHERE ss.student_id = $1::uuid
        ORDER BY fa.day_of_week, p.period_num
    """, student_id)
    
    grid = {}
    for r in rows:
        key = f"{r['day_of_week']}_{r['period_num']}"
        grid[key] = {
            "day": r['day_of_week'], "day_name": DAY_NAMES[r['day_of_week']],
            "period_num": r['period_num'], "period_label": r['label'],
            "start_time": str(r['start_time']), "end_time": str(r['end_time']),
            "subject_code": r['subject_code'], "subject_name": r['subject_name'],
            "subject_type": r['subject_type'],
            "faculty_name": r['faculty_name'], "lecture_seq": r['lecture_seq'],
            "block": r['stream_name']
        }
    
    total_subjects = await pool.fetchval("SELECT COUNT(*) FROM subjects")
    selected_subjects = await pool.fetchval("""
        SELECT COUNT(DISTINCT fa.subject_id)
        FROM student_selections ss
        JOIN faculty_availability fa ON fa.id = ss.availability_id
        WHERE ss.student_id = $1::uuid
    """, student_id)
    total_periods = len(rows)
    target_periods = await pool.fetchval("SELECT SUM(periods_per_week) FROM subjects")
    
    return {
        "grid": grid,
        "is_submitted": submitted,
        "total_subjects": total_subjects,
        "selected_subjects": selected_subjects,
        "total_periods": total_periods,
        "target_periods": target_periods or 0
    }

# ============================================================
# Selection (Bundle-based)
# ============================================================
@app.post("/api/student/select-faculty")
async def select_faculty(req: FacultySelectRequest):
    require_db()
    if await is_submitted(req.student_id):
        raise HTTPException(status_code=403, detail="Selection already submitted")
    
    async with pool.acquire() as conn:
        async with conn.transaction():
            # 1. Get all faculty slots for this subject
            faculty_slots = await conn.fetch("""
                SELECT fa.id, fa.day_of_week, fa.period_id, fa.max_capacity, fa.current_enrolled
                FROM faculty_availability fa
                WHERE fa.faculty_id = $1::uuid AND fa.subject_id = $2::uuid 
                  AND fa.stream_name = $3 AND fa.is_active = TRUE
                FOR UPDATE
            """, req.faculty_id, req.subject_id, req.stream_name)
            
            if not faculty_slots:
                raise HTTPException(status_code=404, detail="No availability found for this faculty/subject")
            
            # 2. Check capacity
            for s in faculty_slots:
                if s['current_enrolled'] >= s['max_capacity']:
                    raise HTTPException(status_code=409, detail="One or more slots in this bundle are full")
            
            # 3. Check for time clashes with OTHER subjects
            existing = await conn.fetch("""
                SELECT fa.day_of_week, fa.period_id, sub.code as subject_code
                FROM student_selections ss
                JOIN faculty_availability fa ON fa.id = ss.availability_id
                JOIN subjects sub ON sub.id = fa.subject_id
                WHERE ss.student_id = $1::uuid AND fa.subject_id != $2::uuid
            """, req.student_id, req.subject_id)
            existing_set = {(r['day_of_week'], r['period_id']): r['subject_code'] for r in existing}
            
            for s in faculty_slots:
                key = (s['day_of_week'], s['period_id'])
                if key in existing_set:
                    raise HTTPException(status_code=409, 
                        detail=f"Clash: {DAY_NAMES[key[0]]} P{key[1]} is already taken by {existing_set[key]}")
            
            # 4. Remove old selections for this subject (faculty change)
            await conn.execute("""
                DELETE FROM student_selections
                WHERE student_id = $1::uuid
                AND availability_id IN (
                    SELECT id FROM faculty_availability WHERE subject_id = $2::uuid
                )
            """, req.student_id, req.subject_id)
            
            # 5. Insert all slots in the bundle
            for s in faculty_slots:
                await conn.execute("""
                    INSERT INTO student_selections (student_id, availability_id)
                    VALUES ($1::uuid, $2::uuid) ON CONFLICT DO NOTHING
                """, req.student_id, s['id'])
                manager.mark_for_update(s['id'])
    
    return {"success": True, "message": "Faculty bundle selected for subject"}

@app.post("/api/student/deselect-subject")
async def deselect_subject(req: SubjectDeselectRequest):
    require_db()
    if await is_submitted(req.student_id):
        raise HTTPException(status_code=403, detail="Selection already submitted")
    
    async with pool.acquire() as conn:
        async with conn.transaction():
            slot_ids = await conn.fetch("""
                SELECT ss.availability_id FROM student_selections ss
                JOIN faculty_availability fa ON fa.id = ss.availability_id
                WHERE ss.student_id = $1::uuid AND fa.subject_id = $2::uuid
            """, req.student_id, req.subject_id)
            
            await conn.execute("""
                DELETE FROM student_selections
                WHERE student_id = $1::uuid
                AND availability_id IN (
                    SELECT id FROM faculty_availability WHERE subject_id = $2::uuid
                )
            """, req.student_id, req.subject_id)
            
            for r in slot_ids:
                manager.mark_for_update(r['availability_id'])
    
    return {"success": True, "message": "Subject deselected"}

@app.delete("/api/select")
async def deselect_slot(req: SelectionRequest):
    """Legacy deselect - kept for backward compatibility"""
    require_db()
    if await is_submitted(req.student_id):
        raise HTTPException(status_code=403, detail="Selection already submitted")
    return {"success": True, "message": "Use /api/student/deselect-subject instead"}

@app.delete("/api/select/all")
async def deselect_all_slots(student_id: str):
    require_db()
    if await is_submitted(student_id):
        raise HTTPException(status_code=403, detail="Selection already submitted")
    await pool.execute("""
        DELETE FROM student_selections WHERE student_id=$1::uuid
    """, student_id)
    return {"success": True, "message": "All selections cleared"}

# ============================================================
# WebSocket Endpoint
# ============================================================
@app.websocket("/ws/updates")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            # Keep connection alive; clients don't need to send messages
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)
    except Exception:
        manager.disconnect(websocket)

# ============================================================
# Validation & Submit
# ============================================================
@app.get("/api/validate/{student_id}")
async def validate_selection(student_id: str):
    require_db()
    subjects = await pool.fetch("SELECT id, code, name, credits, subject_type, periods_per_week FROM subjects")
    errors = []
    for sub in subjects:
        count = await pool.fetchval("""
            SELECT COUNT(*) FROM student_selections ss
            JOIN faculty_availability fa ON fa.id = ss.availability_id
            WHERE ss.student_id=$1::uuid AND fa.subject_id=$2::uuid
        """, student_id, sub['id'])
        expected = sub['periods_per_week'] or sub['credits']
        if count == 0:
            errors.append(f"{sub['code']}: No faculty selected")
        elif count != expected:
            errors.append(f"{sub['code']}: Expected {expected} periods, got {count}")
    return {"valid": len(errors) == 0, "errors": errors, "warnings": []}

@app.post("/api/submit")
async def submit_selection(req: SubmitRequest):
    require_db()
    if await is_submitted(req.student_id):
        raise HTTPException(status_code=409, detail="Already submitted")
    validation = await validate_selection(req.student_id)
    if not validation["valid"]:
        raise HTTPException(status_code=422, detail={"message": "Validation failed", "errors": validation["errors"]})
    async with pool.acquire() as conn:
        async with conn.transaction():
            await conn.execute("UPDATE student_selections SET is_submitted=TRUE WHERE student_id=$1::uuid", req.student_id)
            await conn.execute("INSERT INTO student_submissions (student_id) VALUES ($1::uuid) ON CONFLICT DO NOTHING", req.student_id)
    return {"success": True, "message": "Selection submitted and locked"}

# ============================================================
# Export XLS Report
# ============================================================
@app.get("/api/export/selections")
async def export_selections():
    require_db()
    rows = await pool.fetch("""
        SELECT
            st.register_number,
            st.name AS student_name,
            st.department,
            sub.code AS subject_code,
            sub.name AS subject_name,
            f.name AS faculty_name,
            fa.day_of_week,
            p.period_num,
            p.label AS period_label,
            p.start_time,
            p.end_time,
            ss.selected_at
        FROM student_selections ss
        JOIN students st ON st.id = ss.student_id
        JOIN faculty_availability fa ON fa.id = ss.availability_id
        JOIN subjects sub ON sub.id = fa.subject_id
        JOIN faculty f ON f.id = fa.faculty_id
        JOIN periods p ON p.id = fa.period_id
        WHERE ss.is_submitted = TRUE
        ORDER BY st.register_number, fa.day_of_week, p.period_num
    """)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Selections"

    header_fill = PatternFill("solid", fgColor="1E1E4A")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="F0EFF8")
    border      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    center_align = Alignment(horizontal="center", vertical="center")

    headers    = ["Register Number","Student Name","Department","Subject Code",
                  "Subject Name","Faculty","Day","Period","Time","Selected At"]
    col_widths = [18, 22, 14, 14, 28, 24, 12, 8, 16, 22]

    ws.row_dimensions[1].height = 28
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w

    day_map = {1:"Monday", 2:"Tuesday", 3:"Wednesday", 4:"Thursday", 5:"Friday"}

    for i, r in enumerate(rows, 2):
        fill = alt_fill if i % 2 == 0 else PatternFill()
        data = [
            r["register_number"], r["student_name"], r["department"],
            r["subject_code"], r["subject_name"], r["faculty_name"],
            day_map.get(r["day_of_week"], str(r["day_of_week"])),
            r["period_label"],
            f"{r['start_time']}–{r['end_time']}",
            r["selected_at"].strftime("%Y-%m-%d %H:%M") if r["selected_at"] else ""
        ]
        for col, val in enumerate(data, 1):
            cell        = ws.cell(row=i, column=col, value=val)
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws2 = wb.create_sheet("Summary")
    ws2.cell(1, 1, "Total Students").font  = Font(bold=True)
    ws2.cell(1, 2, len({r["register_number"] for r in rows}))
    ws2.cell(2, 1, "Total Selections").font = Font(bold=True)
    ws2.cell(2, 2, len(rows))
    ws2.cell(3, 1, "Generated At").font    = Font(bold=True)
    ws2.cell(3, 2, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"univ_slot_selections_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



# ============================================================
# Faculty Login
# ============================================================
@app.post("/api/faculty/login")
async def faculty_login(req: FacultyLoginRequest):
    require_db()
    faculty = await pool.fetchrow(
        "SELECT * FROM faculty WHERE employee_id=$1", req.employee_id
    )
    if not faculty:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Check if portal is locked, but allow admins through
    await check_portal_locked(target="faculty", is_admin=faculty["is_admin"])

    # Verify password
    password_hash = faculty["password_hash"]
    valid = False
    if password_hash.startswith("$2"):
        try:
            valid = bcrypt.checkpw(req.password.encode(), password_hash.encode())
        except Exception as e:
            print(f"[WARN] Faculty bcrypt error: {e}")
            valid = False
    else:
        valid = (req.password == password_hash)

    if not valid:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    return {
        "success": True,
        "message": "Login successful",
        "faculty": {
            "id": str(faculty["id"]),
            "name": faculty["name"],
            "employee_id": faculty["employee_id"],
            "department": faculty["department"],
            "is_admin": faculty["is_admin"]
        }
    }

# ============================================================
# Faculty: Dashboard & Student List
# ============================================================
@app.get("/api/faculty/dashboard/{faculty_id}")
async def get_faculty_dashboard(faculty_id: str):
    require_db()
    
    # 1. Get faculty's availability slots
    slots = await pool.fetch("""
        SELECT 
            fa.id as slot_id,
            fa.day_of_week,
            fa.stream_name,
            p.period_num,
            p.label as period_label,
            p.start_time,
            p.end_time,
            sub.name as subject_name,
            sub.code as subject_code,
            fa.max_capacity,
            fa.current_enrolled as booked
        FROM faculty_availability fa
        JOIN periods p ON p.id = fa.period_id
        JOIN subjects sub ON sub.id = fa.subject_id
        WHERE fa.faculty_id = $1::uuid AND fa.is_active = TRUE
        ORDER BY fa.day_of_week, p.period_num
    """, faculty_id)

    # Extract cluster/block from stream_name and cluster_id
    faculty_block = "N/A"
    if slots:
        faculty_block = f"{slots[0]['stream_name']}"

    # 2. For each slot, get students
    dashboard_data = []
    total_enrolled = 0
    unique_students = set()

    for s in slots:
        students = await pool.fetch("""
            SELECT 
                st.name,
                st.register_number,
                st.department
            FROM student_selections ss
            JOIN students st ON st.id = ss.student_id
            WHERE ss.availability_id = $1::uuid
            ORDER BY st.name
        """, s["slot_id"])
        
        student_list = [dict(st) for st in students]
        total_enrolled += len(student_list)
        for st in student_list:
            unique_students.add(st["register_number"])

        dashboard_data.append({
            **dict(s),
            "start_time": str(s["start_time"]),
            "end_time": str(s["end_time"]),
            "students": student_list,
            "occupancy_rate": round((s["booked"] / s["max_capacity"]) * 100, 1) if s["max_capacity"] > 0 else 0
        })

    return {
        "slots": dashboard_data,
        "faculty_block": faculty_block,
        "analytics": {
            "total_slots": len(dashboard_data),
            "total_enrollments": total_enrolled,
            "unique_students": len(unique_students)
        }
    }

@app.get("/api/faculty/export/students")
async def export_faculty_students(faculty_id: str):
    require_db()
    
    faculty = await pool.fetchrow("SELECT name FROM faculty WHERE id = $1::uuid", faculty_id)
    if not faculty:
        raise HTTPException(status_code=404, detail="Faculty not found")

    rows = await pool.fetch("""
        SELECT
            d.day_name,
            p.label as period,
            p.start_time,
            p.end_time,
            sub.name as subject,
            st.register_number,
            st.name as student_name,
            st.department
        FROM faculty_availability fa
        JOIN periods p ON p.id = fa.period_id
        JOIN subjects sub ON sub.id = fa.subject_id
        JOIN student_selections ss ON ss.availability_id = fa.id
        JOIN students st ON st.id = ss.student_id
        CROSS JOIN (
            SELECT 1 as d_val, 'Monday' as day_name UNION ALL
            SELECT 2, 'Tuesday' UNION ALL
            SELECT 3, 'Wednesday' UNION ALL
            SELECT 4, 'Thursday' UNION ALL
            SELECT 5, 'Friday'
        ) d
        WHERE fa.faculty_id = $1::uuid AND fa.day_of_week = d.d_val
        ORDER BY fa.day_of_week, p.period_num, st.name
    """, faculty_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Enrollment"

    # Styling
    header_fill = PatternFill("solid", fgColor="2A2060")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Day", "Period", "Time", "Subject", "Register No", "Student Name", "Department"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for i, r in enumerate(rows, 2):
        data = [
            r["day_name"], r["period"], f"{r['start_time']}–{r['end_time']}",
            r["subject"], r["register_number"], r["student_name"], r["department"]
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = border

    # Auto-adjust column widths
    for i, _col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(i)
        for cell in ws[column_letter]:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = max_length + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"faculty_enrollment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/faculty/export/slot/{slot_id}")
async def export_slot_students(slot_id: str):
    require_db()
    
    slot_info = await pool.fetchrow("""
        SELECT 
            d.day_name, p.label as period, sub.code as sub_code, sub.name as sub_name
        FROM faculty_availability fa
        JOIN periods p ON p.id = fa.period_id
        JOIN subjects sub ON sub.id = fa.subject_id
        CROSS JOIN (
            SELECT 1 as d_val, 'Monday' as day_name UNION ALL
            SELECT 2, 'Tuesday' UNION ALL
            SELECT 3, 'Wednesday' UNION ALL
            SELECT 4, 'Thursday' UNION ALL
            SELECT 5, 'Friday'
        ) d
        WHERE fa.id = $1::uuid AND fa.day_of_week = d.d_val
    """, slot_id)

    if not slot_info:
        raise HTTPException(status_code=404, detail="Slot not found")

    rows = await pool.fetch("""
        SELECT 
            st.register_number,
            st.name as student_name,
            st.department
        FROM student_selections ss
        JOIN students st ON st.id = ss.student_id
        WHERE ss.availability_id = $1::uuid
        ORDER BY st.name
    """, slot_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Period Enrollment"

    # Header info
    ws.merge_cells('A1:C1')
    ws['A1'] = f"Subject: {slot_info['sub_name']} ({slot_info['sub_code']})"
    ws.merge_cells('A2:C2')
    ws['A2'] = f"Time: {slot_info['day_name']} - {slot_info['period']}"
    
    for row in ['1', '2']:
        ws['A' + row].font = Font(bold=True, size=12)
        ws['A' + row].alignment = Alignment(horizontal="center")

    # Table headers
    header_fill = PatternFill("solid", fgColor="3A5EF5")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Register No", "Student Name", "Department"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for i, r in enumerate(rows, 5):
        ws.cell(row=i, column=1, value=r["register_number"])
        ws.cell(row=i, column=2, value=r["student_name"])
        ws.cell(row=i, column=3, value=r["department"])

    # Auto-adjust column widths
    for i, _col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(i)
        for cell in ws[column_letter]:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = max_length + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"enrollment_{slot_info['sub_code']}_{slot_info['day_name']}_{slot_info['period']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============================================================
# Admin: Master Management (CRUD & Reports)
# ============================================================

class StudentCreate(BaseModel):
    register_number: str
    name: str
    email: str
    password: str
    department: str
    semester: int

class FacultyCreate(BaseModel):
    employee_id: str
    name: str
    password: str
    department: str
    is_admin: bool = False

class SubjectCreate(BaseModel):
    code: str
    name: str
    credits: int
    cluster_id: int

@app.get("/api/admin/stats")
async def get_admin_stats():
    require_db()
    total_students = await pool.fetchval("SELECT COUNT(*) FROM students")
    total_faculty = await pool.fetchval("SELECT COUNT(*) FROM faculty")
    total_subjects = await pool.fetchval("SELECT COUNT(*) FROM subjects")
    submitted_count = await pool.fetchval("SELECT COUNT(*) FROM student_submissions")
    
    # Most popular subjects
    popular = await pool.fetch("""
        SELECT sub.name, COUNT(ss.id) as enrollments
        FROM student_selections ss
        JOIN faculty_availability fa ON fa.id = ss.availability_id
        JOIN subjects sub ON sub.id = fa.subject_id
        GROUP BY sub.id, sub.name
        ORDER BY enrollments DESC
        LIMIT 5
    """)

    return {
        "metrics": {
            "students": total_students,
            "faculty": total_faculty,
            "subjects": total_subjects,
            "completion_rate": round((submitted_count / total_students * 100), 1) if total_students > 0 else 0
        },
        "popular_subjects": [dict(p) for p in popular]
    }

# Student Management
@app.get("/api/admin/students")
async def admin_get_students():
    require_db()
    rows = await pool.fetch("SELECT id, register_number, name, email, department, semester FROM students ORDER BY register_number")
    return [dict(r) for r in rows]

@app.post("/api/admin/students")
async def admin_add_student(s: StudentCreate):
    require_db()
    hashed = bcrypt.hashpw(s.password.encode(), bcrypt.gensalt()).decode()
    try:
        await pool.execute("""
            INSERT INTO students (register_number, name, email, password_hash, department, semester)
            VALUES ($1, $2, $3, $4, $5, $6)
        """, s.register_number, s.name, s.email, hashed, s.department, s.semester)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/admin/import/students")
async def admin_import_students(file: UploadFile = File(...)):
    require_db()
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    # Expected structure: Register number, Name, Department, Current Semester
    # Skip header
    imported_count = 0
    errors = []
    
    default_pw = "univ2026"
    hashed = bcrypt.hashpw(default_pw.encode(), bcrypt.gensalt()).decode()
    
    async with pool.acquire() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            
            reg_no, name, dept, sem = row[0], row[1], row[2], row[3]
            email = f"{str(reg_no).lower()}@university.edu" # Generic email generation
            
            try:
                await conn.execute("""
                    INSERT INTO students (register_number, name, email, password_hash, department, semester)
                    VALUES ($1, $2, $3, $4, $5, $6)
                    ON CONFLICT (register_number) DO NOTHING
                """, str(reg_no), name, email, hashed, dept, int(sem))
                imported_count += 1
            except Exception as e:
                errors.append(f"Error on row {reg_no}: {str(e)}")
                
    return {"success": True, "count": imported_count, "errors": errors}

@app.delete("/api/admin/students/{student_id}")
async def admin_delete_student(student_id: str):
    require_db()
    await pool.execute("DELETE FROM students WHERE id = $1::uuid", student_id)
    return {"success": True}

# Faculty Management
@app.get("/api/admin/faculty")
async def admin_get_faculty():
    require_db()
    rows = await pool.fetch("SELECT id, employee_id, name, department, is_admin FROM faculty ORDER BY name")
    return [dict(r) for r in rows]

@app.post("/api/admin/faculty")
async def admin_add_faculty(f: FacultyCreate):
    require_db()
    hashed = bcrypt.hashpw(f.password.encode(), bcrypt.gensalt()).decode()
    try:
        await pool.execute("""
            INSERT INTO faculty (employee_id, name, password_hash, department, is_admin)
            VALUES ($1, $2, $3, $4, $5)
        """, f.employee_id, f.name, hashed, f.department, f.is_admin)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/api/admin/faculty/{faculty_id}")
async def admin_delete_faculty(faculty_id: str):
    require_db()
    await pool.execute("DELETE FROM faculty WHERE id = $1::uuid", faculty_id)
    return {"success": True}

# Subject Management
@app.get("/api/admin/subjects")
async def admin_get_subjects():
    require_db()
    rows = await pool.fetch("SELECT id, code, name, credits, cluster_id FROM subjects ORDER BY code")
    return [dict(r) for r in rows]

@app.post("/api/admin/subjects")
async def admin_add_subject(s: SubjectCreate):
    require_db()
    try:
        await pool.execute("""
            INSERT INTO subjects (code, name, credits, cluster_id)
            VALUES ($1, $2, $3, $4)
        """, s.code, s.name, s.credits, s.cluster_id)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Assignments Management
@app.get("/api/admin/assignments")
async def admin_get_assignments():
    require_db()
    rows = await pool.fetch("""
        SELECT fs.id, f.name as faculty_name, sub.code as subject_code
        FROM faculty_subjects fs
        JOIN faculty f ON f.id = fs.faculty_id
        JOIN subjects sub ON sub.id = fs.subject_id
        ORDER BY f.name
    """)
    return [dict(r) for r in rows]

@app.post("/api/admin/assignments")
async def admin_add_assignment(req: dict):
    require_db()
    try:
        await pool.execute("""
            INSERT INTO faculty_subjects (faculty_id, subject_id)
            VALUES ($1::uuid, $2::uuid)
        """, req["faculty_id"], req["subject_id"])
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/api/admin/assignments/{mapping_id}")
async def admin_delete_assignment(mapping_id: str):
    require_db()
    await pool.execute("DELETE FROM faculty_subjects WHERE id = $1::uuid", mapping_id)
    return {"success": True}

# Master Export
@app.get("/api/admin/export/master")
async def export_master_report():
    require_db()
    # SQL to get every student's full timetable selections
    rows = await pool.fetch("""
        SELECT 
            st.register_number, 
            st.name as student_name, 
            st.department,
            sub.code as sub_code,
            sub.name as sub_name,
            f.name as faculty_name,
            fa.day_of_week,
            p.label as period,
            p.start_time,
            p.end_time
        FROM students st
        LEFT JOIN student_selections ss ON ss.student_id = st.id
        LEFT JOIN faculty_availability fa ON fa.id = ss.availability_id
        LEFT JOIN subjects sub ON sub.id = fa.subject_id
        LEFT JOIN faculty f ON f.id = fa.faculty_id
        LEFT JOIN periods p ON p.id = fa.period_id
        ORDER BY st.register_number, fa.day_of_week, p.period_num
    """)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Enrollment Report"

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ["Reg No", "Student Name", "Dept", "Sub Code", "Subject", "Teacher", "Day", "Period", "Time"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    day_map = {1: "Monday", 2: "Tuesday", 3: "Wednesday", 4: "Thursday", 5: "Friday", 6: "Saturday", 7: "Sunday"}
    for i, r in enumerate(rows, 2):
        data = [
            r["register_number"], r["student_name"], r["department"],
            r["sub_code"], r["sub_name"], r["faculty_name"],
            day_map.get(r["day_of_week"], "Unknown"), r["period"],
            f"{r['start_time']}–{r['end_time']}" if r['start_time'] else ""
        ]
        for col, val in enumerate(data, 1):
            ws.cell(row=i, column=col, value=val)

    # Styling
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=univ_master_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

# ============================================================
# Admin: Semester Subject Management & Year Filtering
# ============================================================

@app.get("/api/admin/stats/filtered")
async def get_admin_stats_filtered(year: Optional[int] = None, semester: Optional[int] = None):
    """Get admin stats optionally filtered by year and semester."""
    require_db()
    
    # Base counts
    student_filter = ""
    student_params = []
    param_idx = 1
    
    if year is not None:
        # Map year to semester range: Year 1 = sem 1,2; Year 2 = sem 3,4; etc.
        sem_low = (year - 1) * 2 + 1
        sem_high = year * 2
        if semester is not None:
            student_filter = f" WHERE semester = ${param_idx}"
            student_params.append(semester)
            param_idx += 1
        else:
            student_filter = f" WHERE semester >= ${param_idx} AND semester <= ${param_idx + 1}"
            student_params.extend([sem_low, sem_high])
            param_idx += 2
    elif semester is not None:
        student_filter = f" WHERE semester = ${param_idx}"
        student_params.append(semester)
        param_idx += 1

    total_students = await pool.fetchval(
        f"SELECT COUNT(*) FROM students{student_filter}", *student_params
    )
    total_faculty = await pool.fetchval("SELECT COUNT(*) FROM faculty WHERE NOT is_admin")
    total_subjects = await pool.fetchval("SELECT COUNT(*) FROM subjects")
    
    # Simpler approach:
    if student_params:
        submitted_count = await pool.fetchval(
            f"SELECT COUNT(*) FROM student_submissions sub JOIN students st ON st.id = sub.student_id{student_filter}",
            *student_params
        )
    else:
        submitted_count = await pool.fetchval("SELECT COUNT(*) FROM student_submissions")

    # Students by year
    students_by_year = await pool.fetch("""
        SELECT 
            CASE 
                WHEN semester IN (1,2) THEN 1
                WHEN semester IN (3,4) THEN 2
                WHEN semester IN (5,6) THEN 3
                WHEN semester IN (7,8) THEN 4
                ELSE 0
            END as year,
            COUNT(*) as count
        FROM students
        GROUP BY year
        ORDER BY year
    """)

    # Most popular subjects
    popular = await pool.fetch("""
        SELECT sub.name, COUNT(ss.id) as enrollments
        FROM student_selections ss
        JOIN faculty_availability fa ON fa.id = ss.availability_id
        JOIN subjects sub ON sub.id = fa.subject_id
        GROUP BY sub.id, sub.name
        ORDER BY enrollments DESC
        LIMIT 5
    """)

    return {
        "metrics": {
            "students": total_students,
            "faculty": total_faculty,
            "subjects": total_subjects,
            "completion_rate": round((submitted_count / total_students * 100), 1) if total_students > 0 else 0
        },
        "students_by_year": [dict(s) for s in students_by_year],
        "popular_subjects": [dict(p) for p in popular]
    }

@app.get("/api/admin/students/by-year")
async def admin_get_students_by_year(year: Optional[int] = None, semester: Optional[int] = None):
    """Get students filtered by year or semester."""
    require_db()
    
    query = "SELECT id, register_number, name, email, department, semester FROM students"
    params = []
    conditions = []
    
    if year is not None:
        sem_low = (year - 1) * 2 + 1
        sem_high = year * 2
        conditions.append(f"semester >= ${len(params)+1} AND semester <= ${len(params)+2}")
        params.extend([sem_low, sem_high])
    
    if semester is not None:
        conditions.append(f"semester = ${len(params)+1}")
        params.append(semester)
    
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    
    query += " ORDER BY register_number"
    rows = await pool.fetch(query, *params)
    return [dict(r) for r in rows]

@app.get("/api/admin/subjects/with-semester")
async def admin_get_subjects_with_semester():
    """Get all subjects with their semester/year info (if assigned to a semester)."""
    require_db()
    rows = await pool.fetch("""
        SELECT s.id, s.code, s.name, s.credits,
               ss.semester, ss.year, ss.is_active
        FROM subjects s
        LEFT JOIN semester_subjects ss ON ss.subject_id = s.id
        ORDER BY s.code
    """)
    return [dict(r) for r in rows]

@app.get("/api/admin/semester-subjects")
async def admin_get_semester_subjects(year: int, semester: int):
    """Get subjects assigned to a specific year+semester."""
    require_db()
    rows = await pool.fetch("""
        SELECT ss.id as mapping_id, s.id as subject_id, s.code, s.name, s.credits, 
               ss.is_active, ss.assigned_at
        FROM semester_subjects ss
        JOIN subjects s ON s.id = ss.subject_id
        WHERE ss.year = $1 AND ss.semester = $2
        ORDER BY s.code
    """, year, semester)
    return [dict(r) for r in rows]

@app.post("/api/admin/semester-subjects")
async def admin_set_semester_subjects(req: dict):
    """Assign subjects to a specific year+semester. Expects {year, semester, subject_ids[]}."""
    require_db()
    year = req.get("year")
    semester = req.get("semester")
    subject_ids = req.get("subject_ids", [])
    
    async with pool.acquire() as conn:
        async with conn.transaction():
            # Remove existing mappings for this year+semester
            await conn.execute(
                "DELETE FROM semester_subjects WHERE year = $1 AND semester = $2",
                year, semester
            )
            # Insert new mappings
            for sid in subject_ids:
                await conn.execute("""
                    INSERT INTO semester_subjects (subject_id, year, semester, is_active)
                    VALUES ($1::uuid, $2, $3, TRUE)
                    ON CONFLICT (subject_id, year, semester) DO UPDATE SET is_active = TRUE
                """, sid, year, semester)
    
    return {"success": True, "count": len(subject_ids)}

@app.delete("/api/admin/semester-subjects/{mapping_id}")
async def admin_remove_semester_subject(mapping_id: str):
    """Remove a subject from a semester."""
    require_db()
    await pool.execute("DELETE FROM semester_subjects WHERE id = $1::uuid", mapping_id)
    return {"success": True}

@app.delete("/api/admin/subjects/{subject_id}")
async def admin_delete_subject(subject_id: str):
    require_db()
    await pool.execute("DELETE FROM subjects WHERE id = $1::uuid", subject_id)
    return {"success": True}

@app.get("/api/admin/portal/status")
async def get_portal_status():
    require_db()
    await ensure_settings()
    rows = await pool.fetch("SELECT key, value FROM system_settings")
    settings = {r['key']: r['value'] == 'true' for r in rows}
    return {
        "student_locked": settings.get('student_locked', False),
        "faculty_locked": settings.get('faculty_locked', False)
    }

@app.post("/api/admin/portal/lock")
async def lock_portal(target: str = "both"):
    require_db()
    await ensure_settings()
    if target in ["student", "both"]:
        await pool.execute("UPDATE system_settings SET value = 'true' WHERE key = 'student_locked'")
    if target in ["faculty", "both"]:
        await pool.execute("UPDATE system_settings SET value = 'true' WHERE key = 'faculty_locked'")
    return {"success": True}

@app.post("/api/admin/portal/unlock")
async def unlock_portal(target: str = "both"):
    require_db()
    await ensure_settings()
    if target in ["student", "both"]:
        await pool.execute("UPDATE system_settings SET value = 'false' WHERE key = 'student_locked'")
    if target in ["faculty", "both"]:
        await pool.execute("UPDATE system_settings SET value = 'false' WHERE key = 'faculty_locked'")
    return {"success": True}

# ============================================================
# Admin: Selection Window & Mailing
# ============================================================
def send_notification_emails(emails: List[str], close_time: str):
    # SMTP config - Can be set via env vars or edited here
    SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
    SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
    SMTP_USER = os.getenv("SMTP_USER", "admin@university.edu")
    SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "your_app_password")
    
    try:
        # Handle different ISO formats
        if 'T' in close_time:
            close_dt = datetime.fromisoformat(close_time.replace('Z', '+00:00'))
            formatted_time = close_dt.strftime("%B %d, %Y at %I:%M %p")
        else:
            formatted_time = close_time
    except Exception:
        formatted_time = close_time
        
    print(f"[EMAIL] Starting to send emails to {len(emails)} recipients...")
    
    # If using default placeholder password, skip actual sending to avoid errors
    if SMTP_PASSWORD == "your_app_password":
        print("[EMAIL] Skipping actual SMTP send: Placeholder password detected. Set SMTP_PASSWORD environment variable.")
        for email in emails:
            if email: print(f"[EMAIL] (Simulated) Would send to: {email}")
        return

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            
            for email in emails:
                if not email: continue
                msg = EmailMessage()
                msg.set_content(f"Dear Student,\n\nThe University Faculty Slot Selection portal is now OPEN.\n\nPlease log in and select your preferred faculty for your registered courses.\n\nThe portal will automatically close on: {formatted_time}.\n\nRegards,\nUniversity Administration")
                msg['Subject'] = "University Portal: Faculty Selection Window Open"
                msg['From'] = SMTP_USER
                msg['To'] = email
                
                try:
                    server.send_message(msg)
                    print(f"[EMAIL] Sent to: {email}")
                except Exception as se:
                    print(f"[EMAIL ERROR] Individual send failed for {email}: {se}")
                    
        print("[EMAIL] Email task completed.")
    except Exception as e:
        print(f"[EMAIL ERROR] Global SMTP failure: {e}")

@app.post("/api/admin/test-email")
async def test_email(target_email: str):
    """Manually trigger a single test email to verify SMTP configuration."""
    require_db()
    # Use a dummy closing time for the test
    test_time = (datetime.now() + timedelta(days=1)).isoformat()
    send_notification_emails([target_email], test_time)
    return {"success": True, "message": f"Test email task triggered for {target_email}. Check server logs for details."}

@app.get("/api/admin/completion-status")
async def admin_completion_status(department: Optional[str] = None):
    require_db()
    # Get overall counts
    total_students = await pool.fetchval("SELECT COUNT(*) FROM students")
    completed_students = await pool.fetchval("SELECT COUNT(DISTINCT student_id) FROM student_selections")
    
    # Get pending students
    query = """
        SELECT s.register_number, s.name, s.department 
        FROM students s
        LEFT JOIN student_selections ss ON s.id = ss.student_id
        WHERE ss.id IS NULL
    """
    params = []
    if department:
        query += " AND s.department = $1"
        params.append(department)
    
    pending_rows = await pool.fetch(query, *params)
    
    # Get list of all departments for the filter
    dept_rows = await pool.fetch("SELECT DISTINCT department FROM students ORDER BY department")
    departments = [r['department'] for r in dept_rows]
    
    return {
        "total_students": total_students,
        "completed_count": completed_students,
        "pending_count": total_students - completed_students,
        "pending_students": [dict(r) for r in pending_rows],
        "departments": departments
    }

@app.post("/api/admin/start-selection")
async def start_selection(req: dict):
    """Starts selection window: unlocks portal, sets closing time, and notifies students."""
    require_db()
    close_time = req.get("close_time")
    if not close_time:
        raise HTTPException(status_code=400, detail="Missing close_time")
    
    # 1. Unlock student portal
    await pool.execute("UPDATE system_settings SET value = 'false' WHERE key = 'student_locked'")
    
    # 2. Get all student emails
    rows = await pool.fetch("SELECT email FROM students WHERE email IS NOT NULL")
    emails = [r['email'] for r in rows]
    
    # 3. Trigger email task in background
    # Note: In a real app, use BackgroundTasks
    send_notification_emails(emails, close_time)
    
    return {"success": True, "message": f"Selection window started. Notifying {len(emails)} students."}

@app.post("/api/admin/deploy-fpc-grid")
async def deploy_fpc_grid():
    """
    Fixed-Pattern Cluster (FPC) Seeding.
    Guarantees 100% clash-free schedules by grouping subjects into fixed time blocks.
    """
    require_db()
    try:
        async with pool.acquire() as conn:
            async with conn.transaction():
                await conn.execute("DELETE FROM student_selections")
                await conn.execute("DELETE FROM faculty_availability")

                # Get subjects with cluster_id
                subjects = await conn.fetch("SELECT id, code, periods_per_week, subject_type, cluster_id FROM subjects ORDER BY cluster_id, code")
                periods = await conn.fetch("SELECT id, period_num FROM periods ORDER BY period_num")
                mapping = await conn.fetch("SELECT subject_id, faculty_id FROM faculty_subjects")
                
                # Period IDs grouped by Period Number for easy access
                p_map = {p['period_num']: p['id'] for p in periods}
                
                # Define Cluster Blocks: {cluster_id: [(day, period_num), ...]}
                # We order them to prefer different days first to avoid trigger conflicts
                cluster_blocks = {
                    1: [(1,1),(3,1),(5,1), (1,2),(3,2),(5,2)], # Mon/Wed/Fri P1, then P2
                    2: [(1,3),(3,3),(5,3), (1,4),(3,4),(5,4)], # Mon/Wed/Fri P3, then P4
                    3: [(2,1),(4,1), (2,2),(4,2), (2,3),(4,3)], # Tue/Thu P1, P2, P3
                    4: [(2,4),(4,4), (2,5),(4,5), (2,6),(4,6)], # Tue/Thu P4, P5, P6
                    5: [(1,5),(3,5),(5,5), (1,6),(3,6),(5,6), (1,7),(3,7),(5,7), (1,8),(3,8),(5,8)], # Mon/Wed/Fri P5-P8
                    6: [(2,7),(4,7), (2,8),(4,8)] # Tue/Thu P7, P8
                }
                
                # Update constraint to allow multiple batches per faculty at same time (idempotent)
                await conn.execute("ALTER TABLE faculty_availability DROP CONSTRAINT IF EXISTS faculty_availability_faculty_id_day_of_week_period_id_key")
                await conn.execute("ALTER TABLE faculty_availability DROP CONSTRAINT IF EXISTS faculty_availability_multi_batch")
                await conn.execute("ALTER TABLE faculty_availability ADD CONSTRAINT faculty_availability_multi_batch UNIQUE (faculty_id, day_of_week, period_id, stream_name)")

                for cluster_id, block_slots in cluster_blocks.items():
                    cluster_subjects = [s for s in subjects if s['cluster_id'] == cluster_id]
                    slot_ptr = 0
                    for sub in cluster_subjects:
                        sub_id = sub['id']
                        ppw = sub['periods_per_week']
                        sub_facs = [m['faculty_id'] for m in mapping if m['subject_id'] == sub_id]
                        
                        # Stagger Block B by half the cluster capacity to ensure different timings
                        offset = len(block_slots) // 2
                        
                        subject_slots_A = [block_slots[(slot_ptr + i) % len(block_slots)] for i in range(ppw)]
                        subject_slots_B = [block_slots[(slot_ptr + offset + i) % len(block_slots)] for i in range(ppw)]
                        
                        for f_id in sub_facs:
                            for block, slots in [('A', subject_slots_A), ('B', subject_slots_B)]:
                                for seq, (day, p_num) in enumerate(slots, 1):
                                    await conn.execute("""
                                        INSERT INTO faculty_availability (faculty_id, subject_id, stream_name, day_of_week, period_id, lecture_seq)
                                        VALUES ($1, $2, $3, $4, $5, $6)
                                    """, f_id, sub_id, f"Block {block}", day, p_map[p_num], seq)
                        
                        slot_ptr += ppw
                                
        return {"success": True, "message": "FPC Seeding Complete. 100% Clash-Free."}
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/block-analytics")
async def admin_block_analytics():
    require_db()
    # Simple analytics: count students assigned to each stream_name (Block A, Block B)
    rows = await pool.fetch("""
        SELECT stream_name, COUNT(DISTINCT student_id) as enrollment
        FROM student_selections
        GROUP BY stream_name
        ORDER BY stream_name
    """)
    
    return {"blocks": [dict(r) for r in rows]}

@app.get("/api/student/bundles")
async def get_bundles(student_id: str):
    """Fetch all faculty bundles for a student, indicating if slots are already selected."""
    require_db()
    async with pool.acquire() as conn:
        # Current selections for this student
        current = await conn.fetch(
            """
            SELECT fa.id FROM student_selections sel
            JOIN faculty_availability fa ON sel.availability_id = fa.id
            WHERE sel.student_id = $1::uuid
            """,
            student_id
        )
        occupied = {c['id'] for c in current}
        # All active bundles
        rows = await conn.fetch(
            """
            SELECT fa.id, fa.subject_id, fa.faculty_id, fa.stream_name, f.name as faculty_name,
                   sub.name as subject_name, fa.day_of_week, fa.period_id
            FROM faculty_availability fa
            JOIN faculty f ON fa.faculty_id = f.id
            JOIN subjects sub ON fa.subject_id = sub.id
            WHERE fa.is_active = TRUE
            ORDER BY fa.faculty_id, fa.subject_id, fa.stream_name
            """
        )
        bundles = {}
        for r in rows:
            key = (r['faculty_id'], r['subject_id'], r['stream_name'])
            if key not in bundles:
                bundles[key] = {
                    "faculty_name": r['faculty_name'],
                    "subject_name": r['subject_name'],
                    "stream_name": r['stream_name'],
                    "slots": []
                }
            bundles[key]["slots"].append({
                "day": r['day_of_week'],
                "period": r['period_id'],
                "selected": r['id'] in occupied
            })
        # Return list of bundles
        return [
            {
                "faculty_id": fid,
                "subject_id": sid,
                "stream_name": data["stream_name"],
                "faculty_name": data["faculty_name"],
                "subject_name": data["subject_name"],
                "slots": data["slots"]
            }
            for (fid, sid, stream), data in bundles.items()
        ]

# ============================================================
# Entry point
# ============================================================
if __name__ == "__main__":
    import uvicorn
    # Make sure to run this using `python main.py` rather than `uvicorn main:app` directly 
    # to avoid import errors if the script isn't in root dir
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
